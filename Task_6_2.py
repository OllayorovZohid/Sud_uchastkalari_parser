import aiohttp
import asyncio
import time
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook

class Uchastka():
    def __init__(self):
        self.title = str()
        self.url = str()
        self.address = str()
        self.contact = str()

async def parse():
    a=[]
    async with aiohttp.ClientSession() as session:
        async with session.get('https://sudrf.ru/index.php?id=300&act=go_ms_search&searchtype=ms&var=true&ms_type=ms&court_subj=0') as response:
            if response.ok:

                # html = await response.text()
                soup = BeautifulSoup(await response.text(),'lxml')
                q = soup.find('table',class_="msSearchResultTbl").find_all('tr')
                try:
                    for tr_tag in q:
                        a.append(tr_tag.find('div',attrs={"style":"padding-top:5px;"}).find('a')['href'])
                    #print(q)
                except:
                    pass
        return a

a_title = []
a_url = []
a_adres = []
a_concat = []

async def parse_problem(session: aiohttp.ClientSession,url):
    # workbook = openpyxl.load_workbook('Task6.xlsx')
    # sheet = workbook['sheet_1']
    Uchastka_1 = Uchastka()
    w=[]
    response = await session.get(url)
    if response.ok:
        try:
            soup = BeautifulSoup(await response.text(), 'lxml')
            Uchastka_1.title = soup.find('p',class_='court-name').text.strip()
            Uchastka_1.url = response.url
            Uchastka_1.address = soup.find('p',attrs={"id":"court_address"}).text.strip()
            #a.append(Uchastka_1.title)
            Uchastka_1.contact = soup.find('p',attrs={"class":"person-phone"}).find('span',class_="right").find('span').text.strip()
            a_title.append(f"{Uchastka_1.title}")
            a_url.append(f"{Uchastka_1.url}")
            a_adres.append(Uchastka_1.address)
            a_concat.append(Uchastka_1.contact)
            print(Uchastka_1.title)
            print(Uchastka_1.url)
            print(Uchastka_1.address)
            print(Uchastka_1.contact)
            # sheet.cell(row_index,3).value = Uchastka_1.title
            # sheet.cell(row_index,4).value = Uchastka_1.url
            # sheet.cell(row_index,5).value = Uchastka_1.address
            # sheet.cell(row_index,6).value = Uchastka_1.contact
            # row_index += 1
        except:
            pass
    # workbook.save('Task6.xlsx')

"""fix yelling at me error"""
from functools import wraps
 
from asyncio.proactor_events import _ProactorBasePipeTransport
 
def silence_event_loop_closed(func):
    @wraps(func)
    def wrapper(self, *args, **kwargs):
        try:
            return func(self, *args, **kwargs)
        except RuntimeError as e:
            if str(e) != 'Event loop is closed':
                raise
    return wrapper
 
_ProactorBasePipeTransport.__del__ = silence_event_loop_closed(_ProactorBasePipeTransport.__del__)
"""fix yelling at me error end"""
q = []

async def main():
    async with aiohttp.ClientSession() as session:
        t1 = time.time()
        tasks = []
        for url in await parse():
            tasks.append(asyncio.create_task(parse_problem(session,url)))

        for task in tasks:
            await task

        
        t2 = time.time()
        print(t2 - t1)
        row_index_1 = 3
        row_index_2 = 3
        row_index_3 = 3
        row_index_4 = 3
        workbook = openpyxl.load_workbook('Task6.xlsx')
        sheet = workbook['sheet_1'] 
        for i_1 in a_url:
            sheet.cell(row_index_1,3).value = i_1
            row_index_1 += 1
        for i_2 in a_title:
            sheet.cell(row_index_2,4).value  = i_2
            row_index_2 += 1
        for i_3 in a_adres:
            sheet.cell(row_index_3,5).value = i_3
            row_index_3 += 1
        #print(a_title)
        for i_4 in a_concat:
            sheet.cell(row_index_4,6).value = i_4
            row_index_4 += 1
    workbook.save('Task6.xlsx')

if __name__ == "__main__":
    asyncio.run(main())